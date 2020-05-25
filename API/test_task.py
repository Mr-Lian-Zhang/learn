#coding:utf-8
import unittest
from API.task import HttpRequest
from openpyxl import load_workbook
wb=load_workbook('zy.xlsx')
sheet=wb['python']
class TestHttp(unittest.TestCase):
    def setUp(self):
        pass
    def test_login_1(self):#正常登陆
        url="https://www.ketangpai.com/UserApi/login"
        data={"email":str(sheet.cell(1,1).value), "password":str(sheet.cell(1,2).value), "remember": 0}
        res=HttpRequest().http_request(url,data,'post')
        try:
            self.assertEqual(1,res.json()["status"])
        except AssertionError as e:
            print("正常登陆的错误是{0}".format(e))
            raise e
    def test_login_2(self):#错误的密码
        url = "https://www.ketangpai.com/UserApi/login"
        data = {"email":str(sheet.cell(2,1).value), "password":str(sheet.cell(2,2).value), "remember": 0}
        res = HttpRequest().http_request(url, data, 'post')
        try:
            self.assertEqual(1,res.json()["status"])
        except AssertionError as e:
            print("错误的密码是{0}".format(e))
            raise e

    def test_login_3(self):  # 账号为空
        url = "https://www.ketangpai.com/UserApi/login"
        data = {"email":str(sheet.cell(3,1).value), "password":str(sheet.cell(3,2).value), "remember": 0}
        res = HttpRequest().http_request(url, data, 'post')
        try:
            self.assertEqual(1,res.json()["status"])
        except AssertionError as e:
            print("账号为空的错误是{0}".format(e))
            raise e
    def test_login_4(self):#密码为空
        url = "https://www.ketangpai.com/UserApi/login"
        data = {"email":str(sheet.cell(4,1).value), "password":str(sheet.cell(4,2).value), "remember": 0}
        res = HttpRequest().http_request(url, data, 'post')
        try:
            self.assertEqual(1,res.json()["status"])
        except AssertionError as e:
            print("密码为空的错误是{0}".format(e))
            raise e

    def tearDown(self):
        pass