#coding:utf-8
#第二种方法！！！·
from openpyxl import load_workbook
class DuExcel:
    def __init__(self,file_name,sheet_name):
        self.file_name=file_name
        self.sheet_name=sheet_name
        self.sheet_obj=load_workbook(self.file_name)[self.sheet_name]
    def get_data(self,i,j):
        return self.sheet_obj.cell(i,j).value
if __name__=='__main__':
    res=DuExcel('zy.xlsx','python').get_data(1,1)
    print(res)