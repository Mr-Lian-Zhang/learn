#coding:utf-8
#如何通过python去操作excel
#1.只支持.xlsx格式 openpyxl
from openpyxl import load_workbook#导入excel模块
wb=load_workbook('zy.xlsx')#指定excel
sheet=wb['python']#指定excel中的表单
res=sheet.cell(1,2).value#拿到表单中指定的值
print(sheet.max_row)#打印 表单中最大行
print(sheet.max_column)#打印表单中最大列
print(res)
#从excel拿出的数据除数字之外其他全是字符串